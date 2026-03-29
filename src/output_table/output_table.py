"""Module to convert dataframes into excel tables"""
import re
import pandas as pd
from openpyxl import load_workbook


class WriteExcel:
    """Class of methods to write to excel"""

    def excel_to_html(self, excel_file, output_html):
        """Converts an Excel file to an HTML file preserving formatting with row numbers.

        Args:
            excel_file (str): The path of the input Excel file.
            output_html (str): The path of the output HTML file.

        Returns:
            str: The path of the generated HTML file.
        """
        wb = load_workbook(excel_file)
        sheet = wb.active
        # Build HTML table with formatting
        html_parts = ['<table border="1" style="border-collapse: collapse;">']
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            html_parts.append('<tr>')
            # Add index column
            if row_idx == 1:
                # Header row - add empty header for index column
                html_parts.append('<th style="background-color: #f2f2f2; font-weight: bold;"></th>')
            else:
                # Data rows - add row number (starting from 0 to match pandas default)
                html_parts.append(f'<td style="font-weight: bold; background-color: #f9f9f9;">{row_idx - 2}</td>')
            for cell in row:
                value = cell.value if cell.value is not None else ''
                value = str(value).replace('\n', '<br>')
                # Build style string based on cell formatting
                styles = []
                # Check if cell has bold formatting
                if cell.font and cell.font.bold:
                    styles.append('font-weight: bold')
                # Check if cell has color formatting
                if cell.font and cell.font.color:
                    # Handle both RGB and theme colors
                    if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                        # RGB color format: 'AARRGGBB' or 'RRGGBB'
                        rgb = cell.font.color.rgb
                        if isinstance(rgb, str):
                            # Remove alpha channel if present (first 2 chars if 8 chars long)
                            if len(rgb) == 8:
                                rgb = rgb[2:]
                            styles.append(f'color: #{rgb}')
                # Determine if this is a header cell
                tag = 'th' if row_idx == 1 else 'td'
                # Build the cell HTML
                if row_idx == 1:
                    # Header cells get default header styling
                    style_attr = ' style="background-color: #f2f2f2; font-weight: bold;"'
                else:
                    style_attr = f' style="{"; ".join(styles)}"' if styles else ''
                html_parts.append(f'<{tag}{style_attr}>{value}</{tag}>')
            html_parts.append('</tr>')
        html_parts.append('</table>')
        # Return only the table fragment (safer for sanitization and embedding)
        html_content = "".join(html_parts)
        with open(output_html, 'w', encoding='utf-8') as file:
            file.write(html_content)
        
        import os
        print(f"DEBUG: Wrote {len(html_content)} bytes to {output_html}")
        return output_html

    def save_doc(self, doc, output_file):
        """Saves docx object to output file"""
        doc.save(output_file)

    def write_excel_acronym_sweep(self, findings, output_file):
        """Write acronym sweep findings to Excel with formatting."""
        # Patterns for underscore markers
        start_marker = re.compile(r'^________ ')   # 8 underscores + space at line start
        end_marker = re.compile(r' _________$')  # space + 9 underscores at line end
        # Color detection (case-insensitive)
        color_pattern = re.compile(
            r'\(\s*(red|green|blue|yellow|pink|violet|teal)\s*\)',
            re.IGNORECASE
        )
        color_map = {
            'red': '#FF0000',
            'green': '#008000',
            'blue': '#0000FF',
            'yellow': '#FFD700',
            'pink': '#FFC0CB',
            'violet': '#8B00FF',
            'teal': '#008080',
        }
        # Process findings to clean text and extract formatting
        processed_data = []
        formatting_info = []
        for finding in findings:
            if pd.isna(finding):
                processed_data.append('')
                formatting_info.append({'bold': False, 'color': None})
                continue
            text = str(finding)
            # Detect presence of markers
            has_start = bool(start_marker.search(text))
            has_end = bool(end_marker.search(text))
            is_bold = has_start or has_end
            # Remove markers from the text
            clean = start_marker.sub('', text)
            clean = end_marker.sub('', clean)
            # Detect color token
            m = color_pattern.search(clean)
            color_hex = color_map[m.group(1).lower()] if m else None
            processed_data.append(clean)
            formatting_info.append({'bold': is_bold, 'color': color_hex})
        # Create DataFrame with processed data
        df = pd.DataFrame({"Acronym Sweep Findings": processed_data})
        # Write Excel with XlsxWriter
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            # Write DataFrame without any data (just structure)
            df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=0)
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            # Create format cache
            format_cache = {}

            def get_format(bold, color):
                key = (bold, color)
                if key not in format_cache:
                    format_dict = {}
                    if bold:
                        format_dict['bold'] = True
                    if color:
                        format_dict['font_color'] = color
                    format_cache[key] = workbook.add_format(format_dict) if format_dict else None
                return format_cache[key]
            # Write header with bold format
            header_format = workbook.add_format({'bold': True})
            worksheet.write(0, 0, "Acronym Sweep Findings", header_format)
            # Write each cell with its specific formatting
            for row_idx, (value, fmt_info) in enumerate(zip(processed_data, formatting_info), start=1):
                cell_format = get_format(fmt_info['bold'], fmt_info['color'])
                if cell_format:
                    worksheet.write(row_idx, 0, value, cell_format)
                else:
                    worksheet.write(row_idx, 0, value)
            # Set column width
            worksheet.set_column(0, 0, 60)
        return output_file
